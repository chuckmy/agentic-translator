"""Process trace log builder and exporters.

Builds a structured trace from a DocumentResult plus run metadata, and exports
it as Markdown or Excel. Also provides bilingual segment-level export (CSV/Excel)
for human review.
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any

from chunker import _split_sentences as _split_target_sentences

PROMPTS = Path(__file__).parent / "prompts"

_XLSX_CELL_MAX = 32000  # below Excel's 32767 hard limit


def _clip(s: Any, n: int = _XLSX_CELL_MAX) -> str:
    s = "" if s is None else str(s)
    return s if len(s) <= n else s[: n - 3] + "..."


def build_trace_data(result, run_meta: dict) -> dict:
    """Flatten DocumentResult into structured trace data for export."""
    chunks_data: list[dict] = []
    iterations_data: list[dict] = []
    errors_data: list[dict] = []

    for ci, (chunk, pr) in enumerate(
        zip(result.chunks, result.chunk_results), start=1
    ):
        stages = pr.stages
        gens = [s for s in stages if s.name == "generation"]
        vers = [s for s in stages if s.name == "verification"]

        chunks_data.append({
            "chunk_index": ci,
            "id_range": chunk.id_range,
            "para_idx": chunk.para_idx,
            "n_segments": len(chunk.segments),
            "source": chunk.text,
            "final_translation": pr.final_translation,
            "iterations": pr.iterations,
            "accepted": pr.accepted,
            "convergence_status": pr.convergence_status,
            "score_history": ",".join(str(s) for s in pr.score_history),
            "tokens_in": sum(s.usage.get("input_tokens", 0) for s in stages),
            "tokens_out": sum(s.usage.get("output_tokens", 0) for s in stages),
            "duration_s": sum(s.duration_s for s in stages),
            "identification": pr.identification,
        })

        for it_idx, (gen, ver) in enumerate(zip(gens, vers), start=1):
            ver_data = ver.output if isinstance(ver.output, dict) else {}
            sc = ver_data.get("spec_compliance") or {}
            iterations_data.append({
                "chunk_index": ci,
                "iteration": it_idx,
                "translation": gen.output if isinstance(gen.output, str) else str(gen.output),
                "mqm_score": ver_data.get("score"),
                "legacy_score": ver_data.get("legacy_score"),
                "penalty": ver_data.get("penalty"),
                "source_length": ver_data.get("source_length"),
                "quality_band": ver_data.get("quality_band", ""),
                "verdict": ver_data.get("verdict"),
                "threshold": ver_data.get("accept_threshold"),
                "n_errors": len(ver_data.get("errors", []) or []),
                "n_concerns": len(sc.get("concerns", []) or []),
                "checked_sections": "; ".join(sc.get("checked_sections", []) or []),
                "concerns": "; ".join(sc.get("concerns", []) or []),
                "summary": ver_data.get("summary", ""),
            })
            for e in ver_data.get("errors", []) or []:
                errors_data.append({
                    "chunk_index": ci,
                    "iteration": it_idx,
                    "severity": e.get("severity", ""),
                    "category": e.get("category", ""),
                    "spec_section": e.get("spec_section", ""),
                    "span": e.get("span", ""),
                    "explanation": e.get("explanation", ""),
                })

    memory_state = {
        "proper_nouns": result.memory.proper_nouns,
        "style_decisions": result.memory.style_decisions,
        "summary": result.memory.summary,
        "chunks_completed": result.memory.chunks_completed,
    }

    return {
        "run_meta": run_meta,
        "chunks": chunks_data,
        "iterations": iterations_data,
        "errors": errors_data,
        "memory_state": memory_state,
        "final_translation": result.final_translation,
    }


def to_markdown(trace: dict) -> str:
    lines: list[str] = []
    rm = trace["run_meta"]

    lines.append("# Agentic Translation — Process Trace")
    lines.append("")
    if rm.get("timestamp"):
        lines.append(f"_Generated at: {rm['timestamp']}_")
        lines.append("")

    lines.append("## Run settings")
    lines.append("")
    for k, v in rm.items():
        if k in ("spec_text", "scoring_criteria"):
            continue
        lines.append(f"- **{k}**: {v}")
    lines.append("")

    criteria = rm.get("scoring_criteria") or {}
    if criteria:
        lines.append("## Scoring criteria")
        lines.append("")
        w = criteria.get("weights", {})
        bands = criteria.get("bands", {})
        lines.append(f"- **Standard**: {criteria.get('standard','')}")
        lines.append(f"- **Normalization**: {criteria.get('normalize_unit','')}")
        lines.append(
            f"- **Severity weights**: critical={w.get('critical')}, "
            f"major={w.get('major')}, minor={w.get('minor')}"
        )
        lines.append("- **Formula**: `score = max(0, 100 − penalty × 100 ÷ source_chars)`")
        lines.append(f"- **Threshold (this run)**: {rm.get('accept_threshold','')}")
        lines.append(
            f"- **Quality bands**: publication ≥ {bands.get('publication_quality','95')}, "
            f"business {bands.get('business_quality','90–94')}, "
            f"needs revision {bands.get('needs_major_revision','<85')}"
        )
        refs = criteria.get("references", [])
        if refs:
            lines.append("- **References**: " + "; ".join(refs))
        lines.append("")

    if rm.get("spec_text"):
        lines.append("## Specification used")
        lines.append("")
        lines.append("```")
        lines.append(str(rm["spec_text"]))
        lines.append("```")
        lines.append("")

    lines.append("## Chunks")
    lines.append("")
    for ch in trace["chunks"]:
        lines.append(f"### Chunk {ch['chunk_index']} ({ch['id_range']})")
        lines.append("")
        lines.append(f"- Iterations: {ch['iterations']} · Status: `{ch['convergence_status']}`")
        if ch["score_history"]:
            lines.append(f"- Score history: {ch['score_history']}")
        lines.append(f"- Tokens: in={ch['tokens_in']} out={ch['tokens_out']} · Duration: {ch['duration_s']:.1f}s")
        lines.append("")
        lines.append("**Source**:")
        lines.append("")
        lines.append("> " + ch["source"].replace("\n", "\n> "))
        lines.append("")

        ident = ch.get("identification") or {}
        devs = ident.get("deviations") or []
        if devs:
            lines.append("**Deviations from Spec defaults**:")
            for d in devs:
                lines.append(
                    f"- [{d.get('dimension','')}] {d.get('observation','')} → {d.get('instruction','')}"
                )
            lines.append("")
        ft = ident.get("focus_terms") or []
        if ft:
            lines.append("**Focus terms**:")
            for f in ft:
                tgt = f.get("tgt", "") or "(unspecified)"
                lines.append(f"- {f.get('src','')} → {tgt}")
            lines.append("")
        if ident.get("notes"):
            lines.append(f"**Chunk-specific notes**: {ident['notes']}")
            lines.append("")

        its = [r for r in trace["iterations"] if r["chunk_index"] == ch["chunk_index"]]
        for it in its:
            lines.append(f"#### Iteration {it['iteration']}")
            lines.append("")
            mqm = it.get("mqm_score")
            band = it.get("quality_band", "")
            mqm_str = f"{mqm:.1f}" if isinstance(mqm, (int, float)) else str(mqm)
            lines.append(
                f"- Verdict: `{it['verdict']}` · MQM: {mqm_str}/100 (band: {band}) · "
                f"Penalty: {it.get('penalty','?')} pts over {it.get('source_length','?')} chars · "
                f"Errors: {it['n_errors']} · Concerns: {it['n_concerns']}"
            )
            if it["checked_sections"]:
                lines.append(f"- Checked sections: {it['checked_sections']}")
            lines.append("")
            lines.append("**Translation**:")
            lines.append("")
            lines.append("> " + (it["translation"] or "").replace("\n", "\n> "))
            lines.append("")
            it_errs = [
                e for e in trace["errors"]
                if e["chunk_index"] == ch["chunk_index"] and e["iteration"] == it["iteration"]
            ]
            if it_errs:
                lines.append("**Errors**:")
                for e in it_errs:
                    sec = f" | Spec: {e['spec_section']}" if e["spec_section"] else ""
                    lines.append(
                        f"- [{(e['severity'] or '').upper()} | {e['category']}{sec}] "
                        f"\"{e['span']}\" — {e['explanation']}"
                    )
                lines.append("")
            if it["concerns"]:
                lines.append(f"**Concerns**: {it['concerns']}")
                lines.append("")

        lines.append("**Accepted translation for this chunk**:")
        lines.append("")
        lines.append("> " + (ch["final_translation"] or "").replace("\n", "\n> "))
        lines.append("")

    lines.append("## Final memory state")
    lines.append("")
    ms = trace["memory_state"]
    if ms["proper_nouns"]:
        lines.append("**Proper-noun ledger**:")
        lines.append("")
        lines.append("| Source | Target |")
        lines.append("|---|---|")
        for s, t in ms["proper_nouns"].items():
            lines.append(f"| {s} | {t} |")
        lines.append("")
    if ms["style_decisions"]:
        lines.append("**Style decisions**:")
        lines.append("")
        lines.append("| Dimension | Decision |")
        lines.append("|---|---|")
        for k, v in ms["style_decisions"].items():
            lines.append(f"| {k} | {v} |")
        lines.append("")
    if ms["summary"]:
        lines.append("**Running summary**:")
        lines.append("")
        lines.append("> " + ms["summary"].replace("\n", "\n> "))
        lines.append("")

    lines.append("## Final translation (full)")
    lines.append("")
    lines.append(trace["final_translation"])
    lines.append("")

    return "\n".join(lines)


def to_excel(trace: dict) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Key", "Value"])
    for k, v in trace["run_meta"].items():
        ws.append([k, _clip(v)])
    ws.append([])
    ws.append(["Final translation", _clip(trace["final_translation"])])

    # Chunks
    ws = wb.create_sheet("Chunks")
    ws.append([
        "chunk_index", "id_range", "para_idx", "n_segments", "iterations",
        "accepted", "convergence_status", "score_history",
        "tokens_in", "tokens_out", "duration_s", "source", "final_translation",
    ])
    for ch in trace["chunks"]:
        ws.append([
            ch["chunk_index"], ch["id_range"], ch["para_idx"], ch["n_segments"],
            ch["iterations"], ch["accepted"], ch["convergence_status"],
            ch["score_history"], ch["tokens_in"], ch["tokens_out"],
            round(ch["duration_s"], 2), _clip(ch["source"]), _clip(ch["final_translation"]),
        ])

    # Iterations
    ws = wb.create_sheet("Iterations")
    ws.append([
        "chunk_index", "iteration", "verdict", "mqm_score", "threshold",
        "penalty", "source_length", "quality_band", "legacy_score",
        "n_errors", "n_concerns", "checked_sections", "concerns",
        "translation", "summary",
    ])
    for it in trace["iterations"]:
        ws.append([
            it["chunk_index"], it["iteration"], it["verdict"],
            it.get("mqm_score"), it.get("threshold"),
            it.get("penalty"), it.get("source_length"), it.get("quality_band", ""),
            it.get("legacy_score"),
            it["n_errors"], it["n_concerns"],
            _clip(it["checked_sections"]), _clip(it["concerns"]),
            _clip(it["translation"]), _clip(it["summary"]),
        ])

    # Errors
    ws = wb.create_sheet("Errors")
    ws.append([
        "chunk_index", "iteration", "severity", "category",
        "spec_section", "span", "explanation",
    ])
    for e in trace["errors"]:
        ws.append([
            e["chunk_index"], e["iteration"], e["severity"], e["category"],
            e["spec_section"], _clip(e["span"]), _clip(e["explanation"]),
        ])

    # Memory
    ws = wb.create_sheet("Memory")
    ws.append(["Type", "Key", "Value"])
    ms = trace["memory_state"]
    for s, t in ms["proper_nouns"].items():
        ws.append(["proper_noun", _clip(s), _clip(t)])
    for k, v in ms["style_decisions"].items():
        ws.append(["style_decision", _clip(k), _clip(v)])
    if ms["summary"]:
        ws.append(["summary", "", _clip(ms["summary"])])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Bilingual segment-level export (for human review)
# ---------------------------------------------------------------------------

_BILINGUAL_HEADERS = [
    "chunk_index", "segment_id", "source", "translation",
    "alignment_quality", "accepted", "score", "n_errors", "notes_for_reviewer",
]


def _extract_json(text: str) -> dict:
    text = (text or "").strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*", "", text).rstrip("`").strip()
    start, end = text.find("{"), text.rfind("}")
    if start == -1 or end == -1:
        raise ValueError(f"No JSON in:\n{text[:300]}")
    return json.loads(text[start : end + 1])


def llm_align_chunk(
    *,
    src_segments,
    target_translation: str,
    target_language: str,
) -> tuple[dict[str, dict], dict]:
    """Call the alignment LLM for one chunk.

    Returns ({segment_id: {target_text, merged, split}}, usage).
    Raises ValueError on parse failure.
    """
    from api import call_model
    system = (PROMPTS / "align_segments.txt").read_text(encoding="utf-8")
    src_block = "\n".join(f"- {s.id}: {s.text}" for s in src_segments)
    user = (
        f"TARGET LANGUAGE: {target_language}\n\n"
        f"SOURCE SEGMENTS:\n{src_block}\n\n"
        f"TARGET TRANSLATION:\n{target_translation}"
    )
    raw, usage = call_model(system=system, user=user, temperature=0.0, max_tokens=2000)
    data = _extract_json(raw)
    aligned = {}
    for a in data.get("alignments", []) or []:
        sid = a.get("segment_id", "")
        if sid:
            aligned[sid] = {
                "target_text": a.get("target_text", ""),
                "merged": bool(a.get("merged", False)),
                "split": bool(a.get("split", False)),
            }
    return aligned, usage


def build_bilingual_rows(
    result,
    *,
    granularity: str = "sentence",  # "sentence" | "chunk"
    target_language: str = "",
    use_llm_for_misaligned: bool = True,
    align_cache: dict | None = None,
) -> list[dict]:
    """Build bilingual rows from a DocumentResult.

    Hybrid alignment strategy when granularity="sentence":
      1. Single-segment chunks → straight row (label: chunk_only).
      2. Multi-segment chunks: try sentence-boundary split on the translation.
         If counts match → 1:1 align (label: segment_regex).
      3. If counts don't match and use_llm_for_misaligned → call LLM aligner
         (label: segment_llm or segment_llm_merged/split per row).
      4. If LLM unavailable / fails → chunk-level row (label: chunk_fallback).

    When granularity="chunk", emit one row per chunk regardless.

    align_cache: optional dict keyed by chunk.index. If passed, LLM results are
    memoised here so repeated downloads don't re-call the LLM.
    """
    rows: list[dict] = []
    for chunk, pr in zip(result.chunks, result.chunk_results):
        src_segs = chunk.segments
        last_score = pr.score_history[-1] if pr.score_history else None
        n_err = len((pr.verification or {}).get("errors", []) or [])

        common = {
            "chunk_index": chunk.index,
            "accepted": pr.accepted,
            "score": last_score,
            "n_errors": n_err,
            "notes_for_reviewer": "",
        }

        if granularity == "chunk":
            rows.append({
                **common,
                "segment_id": chunk.id_range,
                "source": chunk.text,
                "translation": pr.final_translation,
                "alignment_quality": "chunk_level",
            })
            continue

        # granularity == "sentence" below
        if len(src_segs) == 1:
            rows.append({
                **common,
                "segment_id": src_segs[0].id,
                "source": src_segs[0].text,
                "translation": pr.final_translation,
                "alignment_quality": "chunk_only",
            })
            continue

        tgt_sents = _split_target_sentences(pr.final_translation)
        if len(tgt_sents) == len(src_segs):
            for src, tgt in zip(src_segs, tgt_sents):
                rows.append({
                    **common,
                    "segment_id": src.id,
                    "source": src.text,
                    "translation": tgt,
                    "alignment_quality": "segment_regex",
                })
            continue

        # Misaligned — try LLM
        aligned = None
        if use_llm_for_misaligned:
            if align_cache is not None and chunk.index in align_cache:
                aligned = align_cache[chunk.index]
            else:
                try:
                    aligned, _usage = llm_align_chunk(
                        src_segments=src_segs,
                        target_translation=pr.final_translation,
                        target_language=target_language,
                    )
                    if align_cache is not None:
                        align_cache[chunk.index] = aligned
                except Exception:
                    aligned = None

        if aligned and all(s.id in aligned for s in src_segs):
            for src in src_segs:
                a = aligned[src.id]
                q = "segment_llm"
                if a["merged"]:
                    q = "segment_llm_merged"
                elif a["split"]:
                    q = "segment_llm_split"
                rows.append({
                    **common,
                    "segment_id": src.id,
                    "source": src.text,
                    "translation": a["target_text"],
                    "alignment_quality": q,
                })
        else:
            rows.append({
                **common,
                "segment_id": chunk.id_range,
                "source": chunk.text,
                "translation": pr.final_translation,
                "alignment_quality": "chunk_fallback",
            })
    return rows


def bilingual_to_csv(rows: list[dict]) -> bytes:
    """UTF-8-BOM CSV with CRLF line endings (Excel-friendly for Japanese)."""
    buf = io.StringIO(newline="")
    writer = csv.DictWriter(buf, fieldnames=_BILINGUAL_HEADERS, lineterminator="\r\n")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: row.get(h, "") for h in _BILINGUAL_HEADERS})
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def bilingual_to_excel(rows: list[dict]) -> bytes:
    """One-sheet Excel with column widths and wrapped source/translation cells."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Bilingual"
    ws.append(_BILINGUAL_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_clip(row.get(h, "")) for h in _BILINGUAL_HEADERS])

    widths = [8, 12, 60, 60, 18, 10, 8, 10, 40]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    wrap = Alignment(wrap_text=True, vertical="top")
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            if cell.column_letter in ("C", "D"):
                cell.alignment = wrap

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
